#!/usr/bin/env python3
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import subprocess

# 1. data.xlsx 데이터 읽기
df_raw = pd.read_excel('data.xlsx', sheet_name='Sheet1', header=None)

# 헤더는 4행(인덱스 4)에 있음
headers = df_raw.iloc[4].tolist()

# 데이터는 5행(인덱스 5)부터 시작
df_data = df_raw.iloc[5:].copy()
df_data.columns = headers

# 컬럼 매핑 저장 (한 번 선택하면 기억)
column_mapping = {}

# 컬럼명으로 데이터 찾기 (유사한 이름도 찾기)
def get_column_sum(df, column_names, item_name):
    """
    여러 가능한 컬럼명 중 하나를 찾아서 합계 반환
    column_names: 문자열 또는 리스트
    item_name: 항목 이름 (사용자에게 보여줄 이름)
    """
    # 문자열이면 리스트로 변환
    if isinstance(column_names, str):
        column_names = [column_names]
    
    # 이미 매핑된 컬럼이 있으면 사용
    if item_name in column_mapping:
        mapped_col = column_mapping[item_name]
        if mapped_col in df.columns:
            return pd.to_numeric(df[mapped_col], errors='coerce').sum()
        elif mapped_col == 'SKIP':
            return 0
    
    # 각 컬럼명 시도
    for col_name in column_names:
        # 정확히 일치하는 컬럼 찾기
        if col_name in df.columns:
            print(f"✓ '{item_name}' → '{col_name}' 컬럼 사용")
            return pd.to_numeric(df[col_name], errors='coerce').sum()
        
        # 부분 일치하는 컬럼 찾기 (공백, 특수문자 무시)
        for df_col in df.columns:
            if isinstance(df_col, str) and isinstance(col_name, str):
                # 공백 제거하고 비교
                if col_name.replace(' ', '').replace('(', '').replace(')', '') in df_col.replace(' ', '').replace('(', '').replace(')', ''):
                    print(f"✓ '{item_name}' → '{df_col}' 컬럼 사용")
                    return pd.to_numeric(df[df_col], errors='coerce').sum()
    
    # 컬럼을 찾지 못한 경우 사용자에게 선택 요청
    print(f"\n⚠️  '{item_name}' 항목에 해당하는 컬럼을 찾을 수 없습니다.")
    print(f"   찾으려던 컬럼명: {', '.join(column_names)}")
    print(f"\n사용 가능한 컬럼 목록:")
    
    valid_columns = [col for col in df.columns if isinstance(col, str) and col.strip()]
    for idx, col in enumerate(valid_columns, 1):
        print(f"   {idx}. {col}")
    
    print(f"   0. 이 항목 건너뛰기 (0으로 처리)")
    
    while True:
        try:
            choice = input(f"\n'{item_name}' 항목에 사용할 컬럼 번호를 선택하세요 (0-{len(valid_columns)}): ").strip()
            choice_num = int(choice)
            
            if choice_num == 0:
                print(f"→ '{item_name}' 항목을 건너뜁니다.")
                column_mapping[item_name] = 'SKIP'
                return 0
            elif 1 <= choice_num <= len(valid_columns):
                selected_col = valid_columns[choice_num - 1]
                print(f"→ '{item_name}' 항목에 '{selected_col}' 컬럼을 사용합니다.")
                column_mapping[item_name] = selected_col
                return pd.to_numeric(df[selected_col], errors='coerce').sum()
            else:
                print(f"❌ 1부터 {len(valid_columns)} 사이의 숫자를 입력하세요.")
        except ValueError:
            print("❌ 숫자를 입력하세요.")
        except KeyboardInterrupt:
            print("\n\n프로그램을 종료합니다.")
            exit(0)

# 각 항목별 합계 계산 (여러 가능한 컬럼명 제공)
print("\n데이터 컬럼 매칭 중...")
f_column_sum = get_column_sum(df_data, ['바로결제주문금액', '바로결제 주문금액', '직접결제주문금액'], '바로결제주문금액')
g_column_sum = get_column_sum(df_data, ['만나서결제주문금액', '만나서결제 주문금액', '현장결제주문금액'], '만나서결제주문금액')
h_column_sum = get_column_sum(df_data, ['배민1중개이용료', '배민1 중개이용료', '배민 1 중개이용료'], '배민1중개이용료')
i_column_sum = get_column_sum(df_data, ['알뜰배달 중개이용료', '알뜰배달중개이용료'], '알뜰배달 중개이용료')
j_column_sum = get_column_sum(df_data, ['오픈리스트중개이용료', '오픈리스트 중개이용료'], '오픈리스트중개이용료')
k_column_sum = get_column_sum(df_data, ['배민포장주문중개이용료', '배민포장 주문중개이용료', '포장주문중개이용료'], '배민포장주문중개이용료')
l_column_sum = get_column_sum(df_data, ['주문금액 즉시할인', '주문금액즉시할인', '즉시할인'], '주문금액 즉시할인')
m_column_sum = get_column_sum(df_data, ['주문금액 즉시할인 지원', '주문금액즉시할인지원', '즉시할인지원'], '주문금액 즉시할인 지원')
n_column_sum = get_column_sum(df_data, ['바로결제배달팁', '바로결제 배달팁', '직접결제배달팁'], '바로결제배달팁')
o_column_sum = get_column_sum(df_data, ['만나서결제배달팁', '만나서결제 배달팁', '현장결제배달팁'], '만나서결제배달팁')
p_column_sum = get_column_sum(df_data, ['배민클럽(한집배달) 배달팁 할인', '배민클럽한집배달배달팁할인'], '배민클럽(한집배달) 배달팁 할인')
q_column_sum = get_column_sum(df_data, ['배민클럽(한집배달) 배달팁 할인 지원', '배민클럽한집배달배달팁할인지원'], '배민클럽(한집배달) 배달팁 할인 지원')
r_column_sum = get_column_sum(df_data, ['배민클럽(알뜰배달) 배달팁 할인', '배민클럽알뜰배달배달팁할인'], '배민클럽(알뜰배달) 배달팁 할인')
s_column_sum = get_column_sum(df_data, ['배민클럽(알뜰배달) 배달팁 할인 지원', '배민클럽알뜰배달배달팁할인지원'], '배민클럽(알뜰배달) 배달팁 할인 지원')
t_column_sum = get_column_sum(df_data, ['배민1 한집배달 배달비', '배민1한집배달배달비'], '배민1 한집배달 배달비')
u_column_sum = get_column_sum(df_data, ['배민1 한집배달 배달비할인', '배민1한집배달배달비할인'], '배민1 한집배달 배달비할인')
v_column_sum = get_column_sum(df_data, ['알뜰배달 배달비', '알뜰배달배달비'], '알뜰배달 배달비')
w_column_sum = get_column_sum(df_data, ['알뜰배달 배달비할인', '알뜰배달배달비할인'], '알뜰배달 배달비할인')
x_column_sum = get_column_sum(df_data, ['기본수수료(정률)', '기본수수료', '정률수수료'], '기본수수료(정률)')
y_column_sum = get_column_sum(df_data, ['우대수수료', '할인수수료'], '우대수수료')
z_column_sum = get_column_sum(df_data, ['배민 만나서결제주문금액', '배민만나서결제주문금액'], '배민 만나서결제주문금액')
aa_column_sum = get_column_sum(df_data, ['배민 만나서결제배달팁', '배민만나서결제배달팁'], '배민 만나서결제배달팁')
ab_column_sum = get_column_sum(df_data, ['보정금액', '조정금액'], '보정금액')
ac_column_sum = get_column_sum(df_data, ['(E) 부가세', 'E부가세', '부가세E'], '(E) 부가세')
ad_column_sum = get_column_sum(df_data, ['우리가게클릭 이용요금', '우리가게클릭이용요금'], '우리가게클릭 이용요금')
ae_column_sum = get_column_sum(df_data, ['부가세', '부가세F'], '부가세(우리가게클릭)')

print("\n✓ 컬럼 매칭 완료!")

print(f"✓ {len(df_data)}건의 거래 데이터를 읽었습니다.")

# 계산
# 중개이용료 합계 = 배민1 + 알뜰배달 + 오픈리스트 + 배민포장
jungae_total = h_column_sum + i_column_sum + j_column_sum + k_column_sum
# 고객할인 합계 = 주문금액 즉시할인 + 주문금액 즉시할인 지원
gohak_total = l_column_sum + m_column_sum
# 주문중개 합계 = 바로결제주문금액 + 만나서결제주문금액 + 중개이용료 + 고객할인
jumun_jungae_total = f_column_sum + g_column_sum + jungae_total + gohak_total

# 배달비 합계 = 모든 배달비 항목 합
baedalbi_total = n_column_sum + o_column_sum + p_column_sum + q_column_sum + r_column_sum + s_column_sum + t_column_sum + u_column_sum + v_column_sum + w_column_sum

# 결제정산수수료 합계 = 기본수수료 + 우대수수료 + 배민 만나서결제주문금액 + 배민 만나서결제배달팁
gyeoljae_total = x_column_sum + y_column_sum + z_column_sum + aa_column_sum

# 우리가게클릭 합계 = 우리가게클릭 이용요금 + 부가세
urigagae_total = ad_column_sum + ae_column_sum

# (H) 입금금액 = 주문중개 + 배달비 + 결제정산수수료 + 조정금액 + 부가세 + 우리가게클릭
ipgeum_total = jumun_jungae_total + baedalbi_total + gyeoljae_total + ab_column_sum + ac_column_sum + urigagae_total

# 총매출 = 바로결제주문금액 + 만나서결제주문금액 + 바로결제배달팁 + 만나서결제배달팁
total_maechul = f_column_sum + g_column_sum + n_column_sum + o_column_sum

# 매출원가 = 중개이용료 + 고객할인 + 배달비(바로결제/만나서결제 제외) + 기본수수료 + 우대수수료 + 조정금액 + 부가세 + 우리가게클릭
# 배민1중개이용료 + 알뜰배달중개이용료 + 오픈리스트중개이용료 + 배민포장주문중개이용료
# + 주문금액즉시할인 + 주문금액즉시할인지원
# + 배민클럽(한집배달)배달팁할인 + 배민클럽(한집배달)배달팁할인지원
# + 배민클럽(알뜰배달)배달팁할인 + 배민클럽(알뜰배달)배달팁할인지원
# + 배민1한집배달배달비 + 배민1한집배달배달비할인
# + 알뜰배달배달비 + 알뜰배달배달비할인
# + 기본수수료(정률) + 우대수수료
# + 조정금액
# + 부가세
# + 우리가게클릭이용요금 + 부가세(우리가게클릭)
maechul_wonka = (h_column_sum + i_column_sum + j_column_sum + k_column_sum +
                 l_column_sum + m_column_sum +
                 p_column_sum + q_column_sum + r_column_sum + s_column_sum +
                 t_column_sum + u_column_sum + v_column_sum + w_column_sum +
                 x_column_sum + y_column_sum +
                 ab_column_sum +
                 ac_column_sum +
                 ad_column_sum + ae_column_sum)

# 매출총이익 = 총매출 + 매출원가 (매출원가는 음수)
maechul_total_iik = total_maechul + maechul_wonka

print(f"주문중개 합계: {jumun_jungae_total:,.0f}원")
print(f"배달비 합계: {baedalbi_total:,.0f}원")
print(f"결제정산수수료 합계: {gyeoljae_total:,.0f}원")
print(f"조정금액: {ab_column_sum:,.0f}원")
print(f"부가세: {ac_column_sum:,.0f}원")
print(f"우리가게클릭 합계: {urigagae_total:,.0f}원")
print(f"입금금액 합계: {ipgeum_total:,.0f}원")
print(f"\n총매출: {total_maechul:,.0f}원")
print(f"매출원가: {maechul_wonka:,.0f}원")
print(f"매출총이익: {maechul_total_iik:,.0f}원")

# 2. 손익계산서.xlsx 파일 생성 (매번 새로 생성)
output_file = '손익계산서.xlsx'
wb = Workbook()
ws = wb.active
ws.title = "손익계산서"

# 스타일 정의
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
section_fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')

# 제목
ws['A1'] = '손익계산서'
ws['A1'].font = Font(bold=True, size=14)
ws.merge_cells('A1:D1')
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 25

# 헤더
ws['A3'] = '구분'
ws['B3'] = '항목'
ws['C3'] = '금액(원)'
ws['D3'] = '비율(%)'
for col in ['A', 'B', 'C', 'D']:
    cell = ws[f'{col}3']
    cell.font = Font(bold=True, size=11)
    cell.border = border
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Ⅰ. 매출액
ws.merge_cells('A4:B4')
ws['A4'] = 'Ⅰ. 매출액'
ws['A4'].font = Font(bold=True, size=11)
ws['A4'].fill = section_fill
ws['A4'].border = border
ws['A4'].alignment = Alignment(horizontal='left', vertical='center')
ws['C4'] = total_maechul
ws['C4'].number_format = '#,##0'
ws['C4'].font = Font(bold=True, size=11)
ws['C4'].border = border
ws['C4'].fill = section_fill
ws['C4'].alignment = Alignment(horizontal='right', vertical='center')
ws['D4'] = 100.0
ws['D4'].number_format = '0.0'
ws['D4'].font = Font(bold=True, size=11)
ws['D4'].border = border
ws['D4'].fill = section_fill
ws['D4'].alignment = Alignment(horizontal='right', vertical='center')

# (A) 주문중개
ws['A5'] = '(A)'
ws['B5'] = '주문중개'
ws['C5'] = jumun_jungae_total
ws['C5'].number_format = '#,##0'
ws['D5'] = (jumun_jungae_total / total_maechul * 100) if total_maechul != 0 else 0
ws['D5'].number_format = '0.0'
for col in ['A', 'B', 'C', 'D']:
    ws[f'{col}5'].border = border
    ws[f'{col}5'].alignment = Alignment(vertical='center')
ws['C5'].alignment = Alignment(horizontal='right', vertical='center')
ws['D5'].alignment = Alignment(horizontal='right', vertical='center')

# 바로결제주문금액
ws['B6'] = '  바로결제주문금액'
ws['C6'] = f_column_sum
ws['C6'].number_format = '#,##0'
for col in ['A', 'B', 'C']:
    ws[f'{col}6'].border = border
    ws[f'{col}6'].alignment = Alignment(vertical='center')
ws['C6'].alignment = Alignment(horizontal='right', vertical='center')

# 만나서결제주문금액
ws['B7'] = '  만나서결제주문금액'
ws['C7'] = g_column_sum
ws['C7'].number_format = '#,##0'
for col in ['A', 'B', 'C']:
    ws[f'{col}7'].border = border
    ws[f'{col}7'].alignment = Alignment(vertical='center')
ws['C7'].alignment = Alignment(horizontal='right', vertical='center')

# (B) 배달비
ws['A8'] = '(B)'
ws['B8'] = '배달비'
ws['C8'] = baedalbi_total
ws['C8'].number_format = '#,##0'
for col in ['A', 'B', 'C']:
    ws[f'{col}8'].border = border
    ws[f'{col}8'].alignment = Alignment(vertical='center')
ws['C8'].alignment = Alignment(horizontal='right', vertical='center')

# 바로결제배달팁
ws['B9'] = '  바로결제배달팁'
ws['C9'] = n_column_sum
ws['C9'].number_format = '#,##0'
for col in ['A', 'B', 'C']:
    ws[f'{col}9'].border = border
    ws[f'{col}9'].alignment = Alignment(vertical='center')
ws['C9'].alignment = Alignment(horizontal='right', vertical='center')

# 만나서결제배달팁
ws['B10'] = '  만나서결제배달팁'
ws['C10'] = o_column_sum
ws['C10'].number_format = '#,##0'
for col in ['A', 'B', 'C']:
    ws[f'{col}10'].border = border
    ws[f'{col}10'].alignment = Alignment(vertical='center')
ws['C10'].alignment = Alignment(horizontal='right', vertical='center')

# 배민클럽(한집배달) 배달팁 할인
ws['B11'] = '  배민클럽(한집배달) 배달팁 할인'
ws['C11'] = p_column_sum
ws['C11'].number_format = '#,##0'
for col in ['A', 'B', 'C']:
    ws[f'{col}11'].border = border
    ws[f'{col}11'].alignment = Alignment(vertical='center')
ws['C11'].alignment = Alignment(horizontal='right', vertical='center')

# 배민클럽(한집배달) 배달팁 할인 지원
ws['B12'] = '  배민클럽(한집배달) 배달팁 할인 지원'
ws['C12'] = q_column_sum
ws['C12'].number_format = '#,##0'
for col in ['A', 'B', 'C']:
    ws[f'{col}12'].border = border
    ws[f'{col}12'].alignment = Alignment(vertical='center')
ws['C12'].alignment = Alignment(horizontal='right', vertical='center')

# 배민클럽(알뜰배달) 배달팁 할인
ws['B13'] = '  배민클럽(알뜰배달) 배달팁 할인'
ws['C13'] = r_column_sum
ws['C13'].number_format = '#,##0'
for col in ['A', 'B', 'C']:
    ws[f'{col}13'].border = border
    ws[f'{col}13'].alignment = Alignment(vertical='center')
ws['C13'].alignment = Alignment(horizontal='right', vertical='center')

# 배민클럽(알뜰배달) 배달팁 할인 지원
ws['B14'] = '  배민클럽(알뜰배달) 배달팁 할인 지원'
ws['C14'] = s_column_sum
ws['C14'].number_format = '#,##0'
for col in ['A', 'B', 'C']:
    ws[f'{col}14'].border = border
    ws[f'{col}14'].alignment = Alignment(vertical='center')
ws['C14'].alignment = Alignment(horizontal='right', vertical='center')

# 배민1 한집배달 배달비
ws['B15'] = '  배민1 한집배달 배달비'
ws['C15'] = t_column_sum
ws['C15'].number_format = '#,##0'
for col in ['A', 'B', 'C']:
    ws[f'{col}15'].border = border
    ws[f'{col}15'].alignment = Alignment(vertical='center')
ws['C15'].alignment = Alignment(horizontal='right', vertical='center')

# 배민1 한집배달 배달비할인
ws['B16'] = '  배민1 한집배달 배달비할인'
ws['C16'] = u_column_sum
ws['C16'].number_format = '#,##0'
for col in ['A', 'B', 'C']:
    ws[f'{col}16'].border = border
    ws[f'{col}16'].alignment = Alignment(vertical='center')
ws['C16'].alignment = Alignment(horizontal='right', vertical='center')

# 알뜰배달 배달비
ws['B17'] = '  알뜰배달 배달비'
ws['C17'] = v_column_sum
ws['C17'].number_format = '#,##0'
for col in ['A', 'B', 'C']:
    ws[f'{col}17'].border = border
    ws[f'{col}17'].alignment = Alignment(vertical='center')
ws['C17'].alignment = Alignment(horizontal='right', vertical='center')

# 알뜰배달 배달비할인
ws['B18'] = '  알뜰배달 배달비할인'
ws['C18'] = w_column_sum
ws['C18'].number_format = '#,##0'
for col in ['A', 'B', 'C']:
    ws[f'{col}18'].border = border
    ws[f'{col}18'].alignment = Alignment(vertical='center')
ws['C18'].alignment = Alignment(horizontal='right', vertical='center')

# Ⅱ. 매출원가
ws.merge_cells('A20:B20')
ws['A20'] = 'Ⅱ. 매출원가'
ws['A20'].font = Font(bold=True, size=11)
ws['A20'].fill = section_fill
ws['A20'].border = border
ws['A20'].alignment = Alignment(horizontal='left', vertical='center')
ws['C20'].border = border
ws['C20'].fill = section_fill

# 중개이용료
ws['B21'] = '중개이용료'
for col in ['A', 'B', 'C']:
    ws[f'{col}21'].border = border
    ws[f'{col}21'].alignment = Alignment(vertical='center')

# 배민1중개이용료
ws['B22'] = '  배민1중개이용료'
ws['C22'] = h_column_sum
ws['C22'].number_format = '#,##0'
for col in ['A', 'B', 'C']:
    ws[f'{col}22'].border = border
    ws[f'{col}22'].alignment = Alignment(vertical='center')
ws['C22'].alignment = Alignment(horizontal='right', vertical='center')

# 알뜰배달 중개이용료
ws['B23'] = '  알뜰배달 중개이용료'
ws['C23'] = i_column_sum
ws['C23'].number_format = '#,##0'
for col in ['A', 'B', 'C']:
    ws[f'{col}23'].border = border
    ws[f'{col}23'].alignment = Alignment(vertical='center')
ws['C23'].alignment = Alignment(horizontal='right', vertical='center')

# 오픈리스트중개이용료
ws['B24'] = '  오픈리스트중개이용료'
ws['C24'] = j_column_sum
ws['C24'].number_format = '#,##0'
for col in ['A', 'B', 'C']:
    ws[f'{col}24'].border = border
    ws[f'{col}24'].alignment = Alignment(vertical='center')
ws['C24'].alignment = Alignment(horizontal='right', vertical='center')

# 배민포장주문중개이용료
ws['B25'] = '  배민포장주문중개이용료'
ws['C25'] = k_column_sum
ws['C25'].number_format = '#,##0'
for col in ['A', 'B', 'C']:
    ws[f'{col}25'].border = border
    ws[f'{col}25'].alignment = Alignment(vertical='center')
ws['C25'].alignment = Alignment(horizontal='right', vertical='center')

# 고객할인
ws['B27'] = '고객할인'
for col in ['A', 'B', 'C']:
    ws[f'{col}27'].border = border
    ws[f'{col}27'].alignment = Alignment(vertical='center')

# 주문금액 즉시할인
ws['B28'] = '  주문금액 즉시할인'
ws['C28'] = l_column_sum
ws['C28'].number_format = '#,##0'
for col in ['A', 'B', 'C']:
    ws[f'{col}28'].border = border
    ws[f'{col}28'].alignment = Alignment(vertical='center')
ws['C28'].alignment = Alignment(horizontal='right', vertical='center')

# 주문금액 즉시할인 지원
ws['B29'] = '  주문금액 즉시할인 지원'
ws['C29'] = m_column_sum
ws['C29'].number_format = '#,##0'
for col in ['A', 'B', 'C']:
    ws[f'{col}29'].border = border
    ws[f'{col}29'].alignment = Alignment(vertical='center')
ws['C29'].alignment = Alignment(horizontal='right', vertical='center')

# (C) 결제정산수수료
ws['A31'] = '(C)'
ws['B31'] = '결제정산수수료'
ws['C31'] = gyeoljae_total
ws['C31'].number_format = '#,##0'
for col in ['A', 'B', 'C']:
    ws[f'{col}31'].border = border
    ws[f'{col}31'].alignment = Alignment(vertical='center')
ws['C31'].alignment = Alignment(horizontal='right', vertical='center')

# 기본수수료(정률)
ws['B32'] = '  기본수수료(정률)'
ws['C32'] = x_column_sum
ws['C32'].number_format = '#,##0'
for col in ['A', 'B', 'C']:
    ws[f'{col}32'].border = border
    ws[f'{col}32'].alignment = Alignment(vertical='center')
ws['C32'].alignment = Alignment(horizontal='right', vertical='center')

# 우대수수료
ws['B33'] = '  우대수수료'
ws['C33'] = y_column_sum
ws['C33'].number_format = '#,##0'
for col in ['A', 'B', 'C']:
    ws[f'{col}33'].border = border
    ws[f'{col}33'].alignment = Alignment(vertical='center')
ws['C33'].alignment = Alignment(horizontal='right', vertical='center')

# 배민 만나서결제주문금액
ws['B34'] = '  배민 만나서결제주문금액'
ws['C34'] = z_column_sum
ws['C34'].number_format = '#,##0'
for col in ['A', 'B', 'C']:
    ws[f'{col}34'].border = border
    ws[f'{col}34'].alignment = Alignment(vertical='center')
ws['C34'].alignment = Alignment(horizontal='right', vertical='center')

# 배민 만나서결제배달팁
ws['B35'] = '  배민 만나서결제배달팁'
ws['C35'] = aa_column_sum
ws['C35'].number_format = '#,##0'
for col in ['A', 'B', 'C']:
    ws[f'{col}35'].border = border
    ws[f'{col}35'].alignment = Alignment(vertical='center')
ws['C35'].alignment = Alignment(horizontal='right', vertical='center')

# (D) 조정금액
ws['A37'] = '(D)'
ws['B37'] = '조정금액'
ws['C37'] = ab_column_sum
ws['C37'].number_format = '#,##0'
for col in ['A', 'B', 'C']:
    ws[f'{col}37'].border = border
    ws[f'{col}37'].alignment = Alignment(vertical='center')
ws['C37'].alignment = Alignment(horizontal='right', vertical='center')

# (E) 부가세
ws['A39'] = '(E)'
ws['B39'] = '부가세'
ws['C39'] = ac_column_sum
ws['C39'].number_format = '#,##0'
for col in ['A', 'B', 'C']:
    ws[f'{col}39'].border = border
    ws[f'{col}39'].alignment = Alignment(vertical='center')
ws['C39'].alignment = Alignment(horizontal='right', vertical='center')

# (F) 우리가게클릭
ws['A41'] = '(F)'
ws['B41'] = '우리가게클릭'
ws['C41'] = urigagae_total
ws['C41'].number_format = '#,##0'
for col in ['A', 'B', 'C']:
    ws[f'{col}41'].border = border
    ws[f'{col}41'].alignment = Alignment(vertical='center')
ws['C41'].alignment = Alignment(horizontal='right', vertical='center')

# 우리가게클릭 이용요금
ws['B42'] = '  우리가게클릭 이용요금'
ws['C42'] = ad_column_sum
ws['C42'].number_format = '#,##0'
for col in ['A', 'B', 'C']:
    ws[f'{col}42'].border = border
    ws[f'{col}42'].alignment = Alignment(vertical='center')
ws['C42'].alignment = Alignment(horizontal='right', vertical='center')

# 부가세
ws['B43'] = '  부가세'
ws['C43'] = ae_column_sum
ws['C43'].number_format = '#,##0'
for col in ['A', 'B', 'C']:
    ws[f'{col}43'].border = border
    ws[f'{col}43'].alignment = Alignment(vertical='center')
ws['C43'].alignment = Alignment(horizontal='right', vertical='center')

# (H) 입금금액
ws.merge_cells('A45:B45')
ws['A45'] = '(H) 입금금액'
ws['A45'].font = Font(bold=True, size=12, color='FFFFFF')
ws['A45'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
ws['A45'].border = border
ws['A45'].alignment = Alignment(horizontal='center', vertical='center')
ws['C45'] = ipgeum_total
ws['C45'].font = Font(bold=True, size=12)
ws['C45'].number_format = '#,##0'
ws['C45'].border = border
ws['C45'].alignment = Alignment(horizontal='right', vertical='center')
ws['D45'] = (ipgeum_total / total_maechul * 100) if total_maechul != 0 else 0
ws['D45'].font = Font(bold=True, size=12)
ws['D45'].number_format = '0.0'
ws['D45'].border = border
ws['D45'].alignment = Alignment(horizontal='right', vertical='center')

# 빈 행
ws.row_dimensions[46].height = 5

# 총매출
ws.merge_cells('A47:B47')
ws['A47'] = '총매출'
ws['A47'].font = Font(bold=True, size=11)
ws['A47'].fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
ws['A47'].border = border
ws['A47'].alignment = Alignment(horizontal='center', vertical='center')
ws['C47'] = total_maechul
ws['C47'].font = Font(bold=True, size=11)
ws['C47'].number_format = '#,##0'
ws['C47'].border = border
ws['C47'].fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
ws['C47'].alignment = Alignment(horizontal='right', vertical='center')
ws['D47'] = 100.0
ws['D47'].font = Font(bold=True, size=11)
ws['D47'].number_format = '0.0'
ws['D47'].border = border
ws['D47'].fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
ws['D47'].alignment = Alignment(horizontal='right', vertical='center')

# 매출원가
ws.merge_cells('A48:B48')
ws['A48'] = '매출원가'
ws['A48'].font = Font(bold=True, size=11)
ws['A48'].fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
ws['A48'].border = border
ws['A48'].alignment = Alignment(horizontal='center', vertical='center')
ws['C48'] = maechul_wonka
ws['C48'].font = Font(bold=True, size=11)
ws['C48'].number_format = '#,##0'
ws['C48'].border = border
ws['C48'].fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
ws['C48'].alignment = Alignment(horizontal='right', vertical='center')
ws['D48'] = (maechul_wonka / total_maechul * 100) if total_maechul != 0 else 0
ws['D48'].font = Font(bold=True, size=11)
ws['D48'].number_format = '0.0'
ws['D48'].border = border
ws['D48'].fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
ws['D48'].alignment = Alignment(horizontal='right', vertical='center')

# 매출총이익
ws.merge_cells('A49:B49')
ws['A49'] = '매출총이익'
ws['A49'].font = Font(bold=True, size=12, color='FFFFFF')
ws['A49'].fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
ws['A49'].border = border
ws['A49'].alignment = Alignment(horizontal='center', vertical='center')
ws['C49'] = maechul_total_iik
ws['C49'].font = Font(bold=True, size=12, color='FFFFFF')
ws['C49'].number_format = '#,##0'
ws['C49'].border = border
ws['C49'].fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
ws['C49'].alignment = Alignment(horizontal='right', vertical='center')
ws['D49'] = (maechul_total_iik / total_maechul * 100) if total_maechul != 0 else 0
ws['D49'].font = Font(bold=True, size=12, color='FFFFFF')
ws['D49'].number_format = '0.0'
ws['D49'].border = border
ws['D49'].fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
ws['D49'].alignment = Alignment(horizontal='right', vertical='center')

# 모든 데이터 행에 D열 테두리 추가
for row in range(6, 50):
    if ws[f'D{row}'].value is None:
        ws[f'D{row}'].border = border

# 열 너비
ws.column_dimensions['A'].width = 12
ws.column_dimensions['B'].width = 32
ws.column_dimensions['C'].width = 18
ws.column_dimensions['D'].width = 12

# 저장
wb.save(output_file)
print(f"\n✓ {output_file} 파일이 생성되었습니다.")

# 파일 자동으로 열기
subprocess.run(['open', output_file])
