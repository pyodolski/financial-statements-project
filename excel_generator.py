#!/usr/bin/env python3
"""
Excel 손익계산서 생성 모듈
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


def create_styles():
    """Excel 스타일 정의"""
    return {
        'border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        ),
        'header_fill': PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid'),
        'section_fill': PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid'),
        'blue_fill': PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid'),
        'green_fill': PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
    }


def set_cell_style(cell, border, alignment='left', bold=False, size=11, fill=None, color='000000'):
    """셀 스타일 설정"""
    cell.border = border
    cell.font = Font(bold=bold, size=size, color=color)
    cell.alignment = Alignment(horizontal=alignment, vertical='center')
    if fill:
        cell.fill = fill


def add_header(ws, styles):
    """헤더 추가"""
    ws['A1'] = '손익계산서'
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:D1')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 25

    headers = ['구분', '항목', '금액(원)', '비율(%)']
    for idx, header in enumerate(headers, 1):
        col = chr(64 + idx)
        cell = ws[f'{col}3']
        cell.value = header
        set_cell_style(cell, styles['border'], 'center', True, 11, styles['header_fill'])


def add_section_header(ws, row, label, value, ratio, styles, total_maechul):
    """섹션 헤더 추가"""
    ws.merge_cells(f'A{row}:B{row}')
    ws[f'A{row}'] = label
    set_cell_style(ws[f'A{row}'], styles['border'], 'left', True, 11, styles['section_fill'])
    
    ws[f'C{row}'] = value
    ws[f'C{row}'].number_format = '#,##0'
    set_cell_style(ws[f'C{row}'], styles['border'], 'right', True, 11, styles['section_fill'])
    
    ws[f'D{row}'] = ratio if ratio is not None else (value / total_maechul * 100 if total_maechul != 0 else 0)
    ws[f'D{row}'].number_format = '0.0'
    set_cell_style(ws[f'D{row}'], styles['border'], 'right', True, 11, styles['section_fill'])


def add_item_row(ws, row, label, value, styles, show_ratio=False, total_maechul=0, indent=False):
    """항목 행 추가"""
    if label:
        ws[f'A{row}'] = label if not indent else ''
        ws[f'B{row}'] = ('  ' + label) if indent else label
    
    if value is not None:
        ws[f'C{row}'] = value
        ws[f'C{row}'].number_format = '#,##0'
    
    if show_ratio and total_maechul != 0:
        ws[f'D{row}'] = (value / total_maechul * 100) if value else 0
        ws[f'D{row}'].number_format = '0.0'
    
    for col in ['A', 'B', 'C', 'D']:
        if ws[f'{col}{row}'].value is None and col != 'D':
            ws[f'{col}{row}'].value = ''
        ws[f'{col}{row}'].border = styles['border']
        ws[f'{col}{row}'].alignment = Alignment(vertical='center')
    
    ws[f'C{row}'].alignment = Alignment(horizontal='right', vertical='center')
    if show_ratio:
        ws[f'D{row}'].alignment = Alignment(horizontal='right', vertical='center')


def generate_excel(output_file, calc_result):
    """손익계산서 Excel 파일 생성"""
    wb = Workbook()
    ws = wb.active
    ws.title = "손익계산서"
    
    styles = create_styles()
    data = calc_result['data']
    total_maechul = calc_result['total_maechul']
    
    # 헤더
    add_header(ws, styles)
    
    # Ⅰ. 매출액
    add_section_header(ws, 4, 'Ⅰ. 매출액', total_maechul, 100.0, styles, total_maechul)
    
    # (A) 주문중개
    add_item_row(ws, 5, '(A)', None, styles)
    ws['B5'] = '주문중개'
    ws['C5'] = calc_result['jumun_jungae_total']
    ws['C5'].number_format = '#,##0'
    ws['D5'] = (calc_result['jumun_jungae_total'] / total_maechul * 100) if total_maechul != 0 else 0
    ws['D5'].number_format = '0.0'
    for col in ['A', 'B', 'C', 'D']:
        ws[f'{col}5'].border = styles['border']
        ws[f'{col}5'].alignment = Alignment(vertical='center')
    ws['C5'].alignment = Alignment(horizontal='right', vertical='center')
    ws['D5'].alignment = Alignment(horizontal='right', vertical='center')
    
    add_item_row(ws, 6, '바로결제주문금액', data['f'], styles, indent=True)
    add_item_row(ws, 7, '만나서결제주문금액', data['g'], styles, indent=True)
    
    # (B) 배달비
    add_item_row(ws, 8, '(B)', None, styles)
    ws['B8'] = '배달비'
    ws['C8'] = calc_result['baedalbi_total']
    ws['C8'].number_format = '#,##0'
    for col in ['A', 'B', 'C']:
        ws[f'{col}8'].border = styles['border']
        ws[f'{col}8'].alignment = Alignment(vertical='center')
    ws['C8'].alignment = Alignment(horizontal='right', vertical='center')
    
    baedalbi_items = [
        ('바로결제배달팁', data['n']),
        ('만나서결제배달팁', data['o']),
        ('배민클럽(한집배달) 배달팁 할인', data['p']),
        ('배민클럽(한집배달) 배달팁 할인 지원', data['q']),
        ('배민클럽(알뜰배달) 배달팁 할인', data['r']),
        ('배민클럽(알뜰배달) 배달팁 할인 지원', data['s']),
        ('배민1 한집배달 배달비', data['t']),
        ('배민1 한집배달 배달비할인', data['u']),
        ('알뜰배달 배달비', data['v']),
        ('알뜰배달 배달비할인', data['w'])
    ]
    
    row = 9
    for label, value in baedalbi_items:
        add_item_row(ws, row, label, value, styles, indent=True)
        row += 1
    
    # Ⅱ. 매출원가
    row += 1
    ws.merge_cells(f'A{row}:B{row}')
    ws[f'A{row}'] = 'Ⅱ. 매출원가'
    set_cell_style(ws[f'A{row}'], styles['border'], 'left', True, 11, styles['section_fill'])
    ws[f'C{row}'].border = styles['border']
    ws[f'C{row}'].fill = styles['section_fill']
    ws[f'D{row}'].border = styles['border']
    ws[f'D{row}'].fill = styles['section_fill']
    row += 1
    
    # 중개이용료
    add_item_row(ws, row, '', None, styles)
    ws[f'B{row}'] = '중개이용료'
    row += 1
    
    jungae_items = [
        ('배민1중개이용료', data['h']),
        ('알뜰배달 중개이용료', data['i']),
        ('오픈리스트중개이용료', data['j']),
        ('배민포장주문중개이용료', data['k'])
    ]
    
    for label, value in jungae_items:
        add_item_row(ws, row, label, value, styles, indent=True)
        row += 1
    
    row += 1
    
    # 고객할인
    add_item_row(ws, row, '', None, styles)
    ws[f'B{row}'] = '고객할인'
    row += 1
    
    add_item_row(ws, row, '주문금액 즉시할인', data['l'], styles, indent=True)
    row += 1
    add_item_row(ws, row, '주문금액 즉시할인 지원', data['m'], styles, indent=True)
    row += 1
    
    row += 1
    
    # (C) 결제정산수수료
    add_item_row(ws, row, '(C)', None, styles)
    ws[f'B{row}'] = '결제정산수수료'
    ws[f'C{row}'] = calc_result['gyeoljae_total']
    ws[f'C{row}'].number_format = '#,##0'
    for col in ['A', 'B', 'C']:
        ws[f'{col}{row}'].border = styles['border']
        ws[f'{col}{row}'].alignment = Alignment(vertical='center')
    ws[f'C{row}'].alignment = Alignment(horizontal='right', vertical='center')
    row += 1
    
    gyeoljae_items = [
        ('기본수수료(정률)', data['x']),
        ('우대수수료', data['y']),
        ('배민 만나서결제주문금액', data['z']),
        ('배민 만나서결제배달팁', data['aa'])
    ]
    
    for label, value in gyeoljae_items:
        add_item_row(ws, row, label, value, styles, indent=True)
        row += 1
    
    row += 1
    
    # (D) 조정금액
    add_item_row(ws, row, '(D)', None, styles)
    ws[f'B{row}'] = '조정금액'
    ws[f'C{row}'] = data['ab']
    ws[f'C{row}'].number_format = '#,##0'
    for col in ['A', 'B', 'C']:
        ws[f'{col}{row}'].border = styles['border']
        ws[f'{col}{row}'].alignment = Alignment(vertical='center')
    ws[f'C{row}'].alignment = Alignment(horizontal='right', vertical='center')
    row += 1
    
    row += 1
    
    # (E) 부가세
    add_item_row(ws, row, '(E)', None, styles)
    ws[f'B{row}'] = '부가세'
    ws[f'C{row}'] = data['ac']
    ws[f'C{row}'].number_format = '#,##0'
    for col in ['A', 'B', 'C']:
        ws[f'{col}{row}'].border = styles['border']
        ws[f'{col}{row}'].alignment = Alignment(vertical='center')
    ws[f'C{row}'].alignment = Alignment(horizontal='right', vertical='center')
    row += 1
    
    row += 1
    
    # (F) 우리가게클릭
    add_item_row(ws, row, '(F)', None, styles)
    ws[f'B{row}'] = '우리가게클릭'
    ws[f'C{row}'] = calc_result['urigagae_total']
    ws[f'C{row}'].number_format = '#,##0'
    for col in ['A', 'B', 'C']:
        ws[f'{col}{row}'].border = styles['border']
        ws[f'{col}{row}'].alignment = Alignment(vertical='center')
    ws[f'C{row}'].alignment = Alignment(horizontal='right', vertical='center')
    row += 1
    
    add_item_row(ws, row, '우리가게클릭 이용요금', data['ad'], styles, indent=True)
    row += 1
    add_item_row(ws, row, '부가세', data['ae'], styles, indent=True)
    row += 1
    
    row += 1
    
    # (H) 입금금액
    ws.merge_cells(f'A{row}:B{row}')
    ws[f'A{row}'] = '(H) 입금금액'
    ws[f'A{row}'].font = Font(bold=True, size=12, color='FFFFFF')
    ws[f'A{row}'].fill = styles['blue_fill']
    ws[f'A{row}'].border = styles['border']
    ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
    ws[f'C{row}'] = calc_result['ipgeum_total']
    ws[f'C{row}'].font = Font(bold=True, size=12)
    ws[f'C{row}'].number_format = '#,##0'
    ws[f'C{row}'].border = styles['border']
    ws[f'C{row}'].alignment = Alignment(horizontal='right', vertical='center')
    ws[f'D{row}'] = (calc_result['ipgeum_total'] / total_maechul * 100) if total_maechul != 0 else 0
    ws[f'D{row}'].font = Font(bold=True, size=12)
    ws[f'D{row}'].number_format = '0.0'
    ws[f'D{row}'].border = styles['border']
    ws[f'D{row}'].alignment = Alignment(horizontal='right', vertical='center')
    row += 1
    
    # 빈 행
    ws.row_dimensions[row].height = 5
    row += 1
    
    # 총매출
    add_section_header(ws, row, '총매출', total_maechul, 100.0, styles, total_maechul)
    row += 1
    
    # 매출원가
    add_section_header(ws, row, '매출원가', calc_result['maechul_wonka'], 
                      (calc_result['maechul_wonka'] / total_maechul * 100) if total_maechul != 0 else 0,
                      styles, total_maechul)
    row += 1
    
    # 매출총이익
    ws.merge_cells(f'A{row}:B{row}')
    ws[f'A{row}'] = '매출총이익'
    ws[f'A{row}'].font = Font(bold=True, size=12, color='FFFFFF')
    ws[f'A{row}'].fill = styles['green_fill']
    ws[f'A{row}'].border = styles['border']
    ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
    ws[f'C{row}'] = calc_result['maechul_total_iik']
    ws[f'C{row}'].font = Font(bold=True, size=12, color='FFFFFF')
    ws[f'C{row}'].number_format = '#,##0'
    ws[f'C{row}'].border = styles['border']
    ws[f'C{row}'].fill = styles['green_fill']
    ws[f'C{row}'].alignment = Alignment(horizontal='right', vertical='center')
    ws[f'D{row}'] = (calc_result['maechul_total_iik'] / total_maechul * 100) if total_maechul != 0 else 0
    ws[f'D{row}'].font = Font(bold=True, size=12, color='FFFFFF')
    ws[f'D{row}'].number_format = '0.0'
    ws[f'D{row}'].border = styles['border']
    ws[f'D{row}'].fill = styles['green_fill']
    ws[f'D{row}'].alignment = Alignment(horizontal='right', vertical='center')
    
    # D열 테두리 추가
    for r in range(6, row):
        if ws[f'D{r}'].value is None:
            ws[f'D{r}'].border = styles['border']
    
    # 열 너비
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 32
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 12
    
    wb.save(output_file)
